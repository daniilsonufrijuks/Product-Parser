import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# load product names from Excel
wb = openpyxl.load_workbook('products.xlsx')
sheet = wb.active

products = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0]:
        products.append(row[0])

# prepare output workbook
output_wb = openpyxl.Workbook()
output_sheet = output_wb.active
output_sheet.append(["Product", "Shop", "Price", "Product Name"])

# set Selenium
driver = webdriver.Chrome() 
driver.maximize_window()
driver.get("https://www.salidzini.lv/")

for product in products:
    try:
        # wait for search input to be present
        search_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "q"))
        )
        
        # clear and search for product
        search_input.clear()
        search_input.send_keys(product)
        search_input.send_keys(Keys.RETURN)
        
        # wait for results to load
        time.sleep(5)
        
        # scrape results using the correct selectors from the HTML
        items = driver.find_elements(By.CSS_SELECTOR, ".item_box_main")
        
        for item in items[:15]:  # Get top 15 results
            try:
                shop = item.find_element(By.CSS_SELECTOR, ".item_shop_name").text
                price = item.find_element(By.CSS_SELECTOR, ".item_price").text
                product_name = item.find_element(By.CSS_SELECTOR, ".item_name").text
                
                price = price.replace('€', '').strip()
                
                output_sheet.append([product, shop, price, product_name])
            except Exception as e:
                print(f"Error scraping item for '{product}': {str(e)}")
                continue
        
    except Exception as e:
        print(f"Error processing '{product}': {str(e)}")
    
    # ddd small delay 
    time.sleep(5)

# save results
output_wb.save("results.xlsx")
driver.quit()
print("Done! Results saved to 'results.xlsx'.")